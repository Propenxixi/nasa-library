"""
Test untuk validasi urutan kolom batch import
"""
from django.test import TestCase, Client
from django.contrib.auth.models import User
from authentication.models import UserProfile
from django.core.exceptions import ValidationError
from openpyxl import Workbook
import tempfile
import os


class BatchImportValidationTest(TestCase):
    """Test batch import column order validation"""
    
    def setUp(self):
        """Setup test data"""
        # Create a librarian user
        self.librarian = User.objects.create_user(
            username='librarian1',
            password='testpass123'
        )
        UserProfile.objects.create(
            user=self.librarian,
            role='librarian'
        )
        
        self.client = Client()
    
    def create_excel_file(self, headers, data=None):
        """Helper to create Excel file with specified headers and data"""
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write sample data if provided
        if data:
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            return tmp.name
    
    def test_correct_column_order(self):
        """Test that correct column order is accepted"""
        from authentication.views import _validate_excel_columns
        from openpyxl import load_workbook
        
        # Create Excel with correct order
        excel_file = self.create_excel_file(
            ['NIS', 'Nama', 'Jenis Kelamin', 'Kelas'],
            [['2514440', 'Aira Essa', 'P', 'X 1']]
        )
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            result = _validate_excel_columns(ws)
            self.assertTrue(result)
        finally:
            os.unlink(excel_file)
    
    def test_wrong_column_order(self):
        """Test that wrong column order raises ValidationError"""
        from authentication.views import _validate_excel_columns
        from openpyxl import load_workbook
        
        # Create Excel with wrong order: Nama, NIS, Kelas, Jenis Kelamin
        excel_file = self.create_excel_file(
            ['Nama', 'NIS', 'Kelas', 'Jenis Kelamin'],
            [['Aira Essa', '2514440', 'X 1', 'P']]
        )
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            with self.assertRaises(ValidationError) as context:
                _validate_excel_columns(ws)
            
            error_msg = str(context.exception)
            self.assertIn('Urutan kolom tidak sesuai', error_msg)
            self.assertIn('Urutan yang diharapkan', error_msg)
            self.assertIn('Urutan yang ditemukan', error_msg)
        finally:
            os.unlink(excel_file)
    
    def test_missing_column(self):
        """Test that missing column raises ValidationError"""
        from authentication.views import _validate_excel_columns
        from openpyxl import load_workbook
        
        # Missing 'Jenis Kelamin' column
        excel_file = self.create_excel_file(
            ['NIS', 'Nama', 'Kelas'],
            [['2514440', 'Aira Essa', 'X 1']]
        )
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            with self.assertRaises(ValidationError) as context:
                _validate_excel_columns(ws)
            
            error_msg = str(context.exception)
            self.assertIn('tidak ditemukan', error_msg)
            self.assertIn('Jenis Kelamin', error_msg)
        finally:
            os.unlink(excel_file)
    
    def test_extra_columns_with_correct_order(self):
        """Test that extra columns are allowed if required columns are in order"""
        from authentication.views import _validate_excel_columns
        from openpyxl import load_workbook
        
        # Extra columns at the beginning and end
        excel_file = self.create_excel_file(
            ['ID', 'NIS', 'Nama', 'Jenis Kelamin', 'Kelas', 'Telepon'],
            [['1', '2514440', 'Aira Essa', 'P', 'X 1', '081234567']]
        )
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            result = _validate_excel_columns(ws)
            self.assertTrue(result)
        finally:
            os.unlink(excel_file)
    
    def test_columns_out_of_order_complex(self):
        """Test complex order mismatch"""
        from authentication.views import _validate_excel_columns
        from openpyxl import load_workbook
        
        # Order: Jenis Kelamin, Nama, NIS, Kelas
        excel_file = self.create_excel_file(
            ['Jenis Kelamin', 'Nama', 'NIS', 'Kelas'],
            [['P', 'Aira Essa', '2514440', 'X 1']]
        )
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            with self.assertRaises(ValidationError) as context:
                _validate_excel_columns(ws)
            
            error_msg = str(context.exception)
            self.assertIn('Urutan kolom tidak sesuai', error_msg)
        finally:
            os.unlink(excel_file)


if __name__ == '__main__':
    import unittest
    unittest.main()
