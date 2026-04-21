from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from authentication.models import UserProfile
from openpyxl import load_workbook
from django.db import transaction
import os


class Command(BaseCommand):
    help = 'Import students from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            nargs='?',
            type=str,
            help='Path to Excel file (Daftar-Siswa-Cleaned.xlsx)',
        )
        parser.add_argument(
            '--default',
            action='store_true',
            default=False,
            help='Use default path: static/files/Daftar-Siswa-Cleaned.xlsx',
        )
        parser.add_argument(
            '--update',
            action='store_true',
            default=False,
            help='Update existing students (default: skip)',
        )
        parser.add_argument(
            '--delete-first',
            action='store_true',
            default=False,
            help='Delete all existing students before reimporting (fresh import)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Simulate without saving to database',
        )

    def handle(self, *args, **options):
        from django.db import IntegrityError
        from django.contrib.auth.hashers import make_password

        # Determine filepath
        excel_file = options.get('excel_file')
        use_default = options.get('default')
        do_update = options.get('update')
        delete_first = options.get('delete_first')
        dry_run = options.get('dry_run')

        # Handle delete-first flag
        if delete_first and not dry_run:
            student_count = UserProfile.objects.filter(role='student').count()
            if student_count > 0:
                self.stdout.write(self.style.WARNING(f'\n🗑️  Deleting {student_count} existing students...'))
                user_ids = UserProfile.objects.filter(role='student').values_list('user_id', flat=True)
                deleted = User.objects.filter(id__in=user_ids).delete()[0]
                self.stdout.write(self.style.SUCCESS(f'✅ Deleted {deleted} users\n'))

        # Resolve filepath
        if use_default or (not excel_file):
            filepath = 'static/files/Daftar-Siswa-Cleaned.xlsx'
            self.stdout.write(f'📂 Using default path: {filepath}\n')
        else:
            filepath = excel_file

        # Check if file exists
        if not os.path.exists(filepath):
            raise CommandError(f'File not found: {filepath}')

        self.stdout.write(f'📂 Reading file: {filepath}\n')
        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  DRY RUN — no data will be saved\n'))

        try:
            workbook = load_workbook(filepath, data_only=True)
            worksheet = workbook.active

            imported_count = 0
            skipped_count = 0
            updated_count = 0
            error_count = 0

            self.stdout.write('─' * 60)
            self.stdout.write('📖 Reading and preparing data...\n')

            # First pass: read all data and prepare
            all_readied_data = []

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Handle row safely with bounds checking
                    if len(row) < 2:
                        skipped_count += 1
                        continue

                    nis = row[0]
                    nama = row[1]
                    jenis_kelamin = row[2] if len(row) > 2 else None
                    kelas = row[3] if len(row) > 3 else None

                    # Validate data
                    if not nis or not nama:
                        skipped_count += 1
                        continue

                    # Convert to string and strip whitespace
                    if isinstance(nis, float):
                        nis = str(int(nis))
                    else:
                        nis = str(nis).strip()
                        if nis.endswith('.0'):
                            nis = nis[:-2]

                    nama = str(nama).strip()
                    jenis_kelamin = str(jenis_kelamin).strip() if jenis_kelamin else ''
                    kelas = str(kelas).strip() if kelas else ''

                    # Check if user already exists
                    try:
                        existing_user = User.objects.get(username=nis)
                    except User.DoesNotExist:
                        existing_user = None

                    if existing_user:
                        if do_update:
                            # Will update below
                            all_readied_data.append(('update', existing_user, nis, nama, jenis_kelamin, kelas))
                        else:
                            skipped_count += 1
                        continue

                    # Store for bulk create
                    all_readied_data.append(('create', None, nis, nama, jenis_kelamin, kelas))
                    imported_count += 1

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'  Row {row_idx}: ❌ Read error {e.__class__.__name__}'))
                    error_count += 1

            workbook.close()

            if not all_readied_data:
                self.stdout.write(self.style.WARNING('No data to import or update'))
                return

            self.stdout.write(f'✅ Prepared {len(all_readied_data)} records\n')

            # Process updates first (faster)
            updates = [d for d in all_readied_data if d[0] == 'update']
            creates = [d for d in all_readied_data if d[0] == 'create']

            if updates:
                self.stdout.write(f'📝 Updating {len(updates)} existing students...\n')
                with transaction.atomic():
                    for op_type, existing_user, nis, nama, jenis_kelamin, kelas in updates:
                        parts = nama.split(' ', 1)
                        existing_user.first_name = parts[0]
                        existing_user.last_name = parts[1] if len(parts) > 1 else ''
                        existing_user.save()

                        profile = existing_user.profile
                        profile.gender = jenis_kelamin
                        profile.kelas = kelas
                        profile.save()
                    self.stdout.write(self.style.SUCCESS(f'✅ Updated {len(updates)} students\n'))
                    updated_count = len(updates)

            # Process creates in chunks with progress
            if creates:
                self.stdout.write(f'📝 Creating {len(creates)} new students...\n')
                chunk_size = 100
                users_to_create = []
                profiles_to_create = []

                with transaction.atomic():
                    for i, (op_type, _, nis, nama, jenis_kelamin, kelas) in enumerate(creates):
                        parts = nama.split(' ', 1)
                        first_name = parts[0]
                        last_name = parts[1] if len(parts) > 1 else ''

                        # Hash password with NIS (username and password are the same - NIS)
                        password_field = make_password(nis)

                        user = User(
                            username=nis,
                            password=password_field,
                            first_name=first_name,
                            last_name=last_name,
                            email=f'{nis}@student.local'
                        )
                        users_to_create.append(user)

                        profiles_to_create.append({
                            'user': user,
                            'nis': nis,
                            'gender': jenis_kelamin,
                            'kelas': kelas,
                            'role': 'student'
                        })

                        # Process in chunks
                        if len(users_to_create) >= chunk_size:
                            self.stdout.write(f'  Creating chunk {i // chunk_size + 1}... {i+1}/{len(creates)} rows', ending='\r')

                            created_users = User.objects.bulk_create(users_to_create, ignore_conflicts=False)

                            profiles = [
                                UserProfile(
                                    user=profile_data['user'],
                                    nis=profile_data['nis'],
                                    gender=profile_data['gender'],
                                    kelas=profile_data['kelas'],
                                    role=profile_data['role']
                                )
                                for profile_data in profiles_to_create
                            ]
                            UserProfile.objects.bulk_create(profiles, ignore_conflicts=False)

                            users_to_create = []
                            profiles_to_create = []

                    # Process remaining users
                    if users_to_create:
                        self.stdout.write(f'  Creating final chunk... {len(creates)}/{len(creates)}', ending='')
                        created_users = User.objects.bulk_create(users_to_create, ignore_conflicts=False)

                        profiles = [
                            UserProfile(
                                user=profile_data['user'],
                                nis=profile_data['nis'],
                                gender=profile_data['gender'],
                                kelas=profile_data['kelas'],
                                role=profile_data['role']
                            )
                            for profile_data in profiles_to_create
                        ]
                        UserProfile.objects.bulk_create(profiles, ignore_conflicts=False)

                self.stdout.write(self.style.SUCCESS(f'\n✅ Created {len(creates)} students\n'))

            # Print summary
            self.stdout.write('\n' + '═' * 60)
            self.stdout.write(self.style.SUCCESS(
                f'🎉 SELESAI\n'
                f'   ✅ Imported : {imported_count}\n'
                f'   ✏️  Updated  : {updated_count}\n'
                f'   ⏭  Skipped  : {skipped_count}\n'
                f'   ❌ Errors   : {error_count}\n'
                f'   📊 Total   : {imported_count + updated_count + skipped_count + error_count}'
            ))

        except Exception as e:
            raise CommandError(f'Error: {str(e)}')
