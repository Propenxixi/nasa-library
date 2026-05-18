from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.core.exceptions import PermissionDenied
from django.core.exceptions import ValidationError
from django.http import StreamingHttpResponse
import tempfile
import os
import json
from openpyxl import load_workbook
from .forms import LoginForm, CustomPasswordChangeForm, ChangeUsernameForm, StudentBatchImportForm, ProfileUpdateForm
from .models import UserProfile


@require_http_methods(["GET", "POST"])
def user_login(request):
    """Login view - allow both authenticated and unauthenticated users"""
    
    if request.user.is_authenticated:
        return redirect('main:mainpage')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            nis = form.cleaned_data['nis']
            password = form.cleaned_data['password']
            
            try:
                user = User.objects.get(username=nis)
                
                # Check if student is deactivated (soft deleted)
                if hasattr(user, 'profile') and user.profile.is_student():
                    if not user.profile.is_active_student:
                        reason = user.profile.get_deactivation_reason_display() or 'Tidak diketahui'
                        messages.error(
                            request, 
                            f'Akun Anda telah dinonaktifkan. Alasan: {reason}. Hubungi petugas perpustakaan untuk informasi lebih lanjut.'
                        )
                        return render(request, 'login.html', {'form': form})
                
                # Authenticate user
                user = authenticate(request, username=nis, password=password)
                
                if user is not None:
                    login(request, user)
                    messages.success(request, f'Selamat datang, {user.first_name}!')
                    return redirect('main:mainpage')
                else:
                    messages.error(request, 'Password salah. Silahkan coba lagi.')
            except User.DoesNotExist:
                messages.error(request, 'Username atau NIS tidak ditemukan')
        else:
            messages.error(request, 'Silahkan isi semua field dengan benar.')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})


@require_http_methods(["GET", "POST"])
@login_required(login_url='authentication:login')
def change_password(request):
    """Change password view"""
    
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Password Anda berhasil diubah!')
            return redirect('main:mainpage')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CustomPasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'page_title': 'Ubah Password'
    }
    return render(request, 'change_password.html', context)


@require_http_methods(["GET", "POST"])
@login_required(login_url='authentication:login')
def change_username(request):
    """Change username view"""

    if request.method == 'POST':
        form = ChangeUsernameForm(request.POST)
        if form.is_valid():
            new_username = form.cleaned_data['new_username']
            password = form.cleaned_data['password']

            user = authenticate(request, username=request.user.username, password=password)

            if user is not None:
                user.username = new_username
                user.save()
                messages.success(request, 'Username Anda berhasil diubah!')
                return redirect('main:mainpage')
            else:
                messages.error(request, 'Password yang Anda masukkan salah.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = ChangeUsernameForm()

    context = {
        'form': form,
        'page_title': 'Ubah Username'
    }
    return render(request, 'change_username.html', context)


@require_http_methods(["GET", "POST"])
@login_required(login_url='authentication:login')
def profile(request):
    """User profile management view - edit username, password, and profile picture"""

    user_profile = request.user.profile

    form_errors = None
    
    if request.method == 'POST':
        form = ProfileUpdateForm(request.user, request.POST, request.FILES, instance=user_profile)
        if form.is_valid():
            # Check if there are actual changes
            if not form.has_changes():
                form_errors = ['Tidak ada perubahan yang dibuat. Silakan isi minimal satu bidang untuk diubah.']
            else:
                form.save()
                messages.success(request, 'Profil Anda berhasil diperbarui!')
                return redirect('authentication:profile')
        else:
            form_errors = []
            for field, errors in form.errors.items():
                for error in errors:
                    form_errors.append(str(error))
    else:
        form = ProfileUpdateForm(request.user, instance=user_profile)

    context = {
        'form': form,
        'user_profile': user_profile,
        'page_title': 'Kelola Profil',
        'form_errors': json.dumps(form_errors) if form_errors else None
    }
    return render(request, 'profile.html', context)


@require_http_methods(["POST"])
@login_required(login_url='authentication:login')
def user_logout(request):
    """Logout view"""
    # Clear all pending messages before logging out
    storage = messages.get_messages(request)
    for _ in storage:
        pass  # Consume the iterator to clear messages
    logout(request)
    messages.success(request, 'Anda telah berhasil keluar.')
    return redirect('authentication:login')


def _is_librarian(user):
    """Check if user is librarian or superuser"""
    if user.is_superuser:
        return True
    return hasattr(user, "profile") and user.profile.is_librarian()


def _validate_excel_columns(worksheet):
    """Validate that Excel file has required columns and correct order"""
    required_columns = ['NIS', 'Nama', 'Jenis Kelamin', 'Kelas']
    
    # Read header row
    header_row = []
    for cell in worksheet[1]:
        if cell.value:
            header_row.append(str(cell.value).strip())
    
    # Check if required columns exist
    required_set = set(required_columns)
    header_set = set(header_row)
    missing_columns = required_set - header_set
    
    if missing_columns:
        raise ValidationError(
            f'Kolom berikut tidak ditemukan: {", ".join(sorted(missing_columns))}. '
            f'Kolom yang diperlukan: {", ".join(required_columns)}'
        )
    
    # Check if columns are in the correct order
    # Get indices of required columns
    required_indices = []
    for col_name in required_columns:
        try:
            idx = header_row.index(col_name)
            required_indices.append(idx)
        except ValueError:
            pass
    
    # Check if required columns are in order (indices should be increasing)
    if required_indices != sorted(required_indices):
        actual_order = [header_row[idx] for idx in sorted(required_indices)]
        raise ValidationError(
            f'❌ Urutan kolom tidak sesuai!\n\n'
            f'Urutan yang diharapkan:\n'
            f'1. NIS\n'
            f'2. Nama\n'
            f'3. Jenis Kelamin\n'
            f'4. Kelas\n\n'
            f'Urutan yang ditemukan:\n' + 
            '\n'.join([f'{i+1}. {col}' for i, col in enumerate(actual_order)])
        )
    
    return True


def _process_import_stream(tmp_file_path, update_existing):
    """
    Generator function untuk proses import dan yield progress updates.
    
    Proses:
    1. Import/update siswa dari Excel
    2. Setelah selesai, soft delete siswa yang tidak ada di Excel (diasumsikan lulus)
    """
    
    try:
        from django.utils import timezone
        
        # Load workbook
        workbook = load_workbook(tmp_file_path, data_only=True)
        worksheet = workbook.active
        
        # Validate columns
        _validate_excel_columns(worksheet)
        
        # Get column indices
        header_row = list(worksheet[1])
        col_map = {}
        for idx, cell in enumerate(header_row):
            if cell.value:
                col_name = str(cell.value).strip()
                col_map[col_name] = idx
        
        # Count total rows
        total_rows = worksheet.max_row - 1  # Exclude header
        
        yield json.dumps({
            'type': 'started',
            'total_rows': total_rows,
            'message': f'Memulai import {total_rows} siswa...'
        }) + '\n'
        
        imported_count = 0
        skipped_count = 0
        updated_count = 0
        error_count = 0
        deactivated_count = 0
        reactivated_count = 0
        
        # Track NIS yang ada di Excel
        excel_nis_set = set()
        
        # Phase 1: Process rows dari Excel
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                # Extract values from mapped columns
                nis_cell = row[col_map['NIS']]
                nama_cell = row[col_map['Nama']]
                gender_cell = row[col_map['Jenis Kelamin']]
                kelas_cell = row[col_map['Kelas']]
                
                # NIS is a whole number (integer). Handle float from Excel and convert to string
                try:
                    nis = str(int(float(nis_cell.value))) if nis_cell.value else None
                except (ValueError, TypeError):
                    nis = None
                
                nama = str(nama_cell.value).strip() if nama_cell.value else None
                gender = str(gender_cell.value).strip() if gender_cell.value else None
                kelas = str(kelas_cell.value).strip() if kelas_cell.value else None
                
                # Validate required fields
                if not nis or not nama or not gender or not kelas:
                    skipped_count += 1
                    yield json.dumps({
                        'type': 'row',
                        'row': row_idx,
                        'status': 'skipped',
                        'nis': nis or '?',
                        'nama': nama or '-',
                        'reason': 'Data tidak lengkap'
                    }) + '\n'
                    continue
                
                # Validate gender value (L or P)
                if gender not in ['L', 'P']:
                    error_count += 1
                    skipped_count += 1
                    yield json.dumps({
                        'type': 'row',
                        'row': row_idx,
                        'status': 'error',
                        'nis': nis,
                        'nama': nama,
                        'reason': f"Jenis Kelamin harus 'L' atau 'P' (ditemukan: {gender})"
                    }) + '\n'
                    continue
                
                # Add to set of valid NIS
                excel_nis_set.add(nis)
                
                # Split nama into first_name and last_name
                nama_parts = nama.split(maxsplit=1)
                first_name = nama_parts[0]
                last_name = nama_parts[1] if len(nama_parts) > 1 else ''
                
                # Check if student already exists
                try:
                    user = User.objects.get(username=nis)
                    profile = user.profile
                    
                    if update_existing:
                        # Check if student was previously deactivated, reactivate if so
                        was_deactivated = not profile.is_active_student
                        
                        # Update existing user
                        user.first_name = first_name
                        user.last_name = last_name
                        user.is_active = True  # Ensure active
                        user.save()
                        
                        # Update profile
                        profile.gender = gender
                        profile.kelas = kelas
                        profile.is_active_student = True
                        profile.deactivated_at = None
                        profile.deactivation_reason = None
                        profile.save()
                        
                        if was_deactivated:
                            reactivated_count += 1
                            yield json.dumps({
                                'type': 'row',
                                'row': row_idx,
                                'status': 'updated',
                                'nis': nis,
                                'nama': nama,
                                'reason': 'Siswa reaktifkan (sebelumnya nonaktif)'
                            }) + '\n'
                        else:
                            updated_count += 1
                            yield json.dumps({
                                'type': 'row',
                                'row': row_idx,
                                'status': 'updated',
                                'nis': nis,
                                'nama': nama,
                                'reason': 'Data diperbarui'
                            }) + '\n'
                    else:
                        skipped_count += 1
                        yield json.dumps({
                            'type': 'row',
                            'row': row_idx,
                            'status': 'skipped',
                            'nis': nis,
                            'nama': nama,
                            'reason': 'Siswa sudah ada (skip)'
                        }) + '\n'
                
                except User.DoesNotExist:
                    # Create new user
                    # Username dan password default = NIS
                    user = User.objects.create_user(
                        username=nis,
                        first_name=first_name,
                        last_name=last_name,
                        password=nis,  # Default password = NIS
                        is_active=True
                    )
                    
                    # Create user profile
                    UserProfile.objects.create(
                        user=user,
                        role='student',
                        nis=nis,
                        gender=gender,
                        kelas=kelas,
                        is_active_student=True
                    )
                    
                    imported_count += 1
                    yield json.dumps({
                        'type': 'row',
                        'row': row_idx,
                        'status': 'success',
                        'nis': nis,
                        'nama': nama,
                        'reason': 'Siswa baru ditambahkan'
                    }) + '\n'
            
            except Exception as e:
                error_count += 1
                skipped_count += 1
                yield json.dumps({
                    'type': 'row',
                    'row': row_idx,
                    'status': 'error',
                    'nis': '?',
                    'nama': '?',
                    'reason': str(e)
                }) + '\n'
        
        # Phase 2: Soft delete siswa yang tidak ada di Excel (diasumsikan lulus)
        # Only soft delete active students
        try:
            students_to_deactivate = UserProfile.objects.filter(
                role='student',
                is_active_student=True
            ).exclude(nis__in=excel_nis_set)
            
            for profile in students_to_deactivate:
                profile.deactivate(reason='graduated')
                deactivated_count += 1
                
                yield json.dumps({
                    'type': 'row',
                    'row': 0,
                    'status': 'deactivated',
                    'nis': profile.nis,
                    'nama': f"{profile.user.first_name} {profile.user.last_name}",
                    'reason': 'Diasumsikan lulus (tidak ada di file baru)'
                }) + '\n'
        except Exception as e:
            yield json.dumps({
                'type': 'error',
                'message': f'Kesalahan saat deactivasi siswa: {str(e)}'
            }) + '\n'
        
        # Final summary
        yield json.dumps({
            'type': 'completed',
            'imported_count': imported_count,
            'updated_count': updated_count,
            'reactivated_count': reactivated_count,
            'deactivated_count': deactivated_count,
            'skipped_count': skipped_count,
            'error_count': error_count,
            'message': f'Import selesai: {imported_count} ditambahkan, {updated_count} diperbarui, {reactivated_count} reaktifkan, {deactivated_count} nonaktifkan, {error_count} error'
        }) + '\n'
    
    except ValidationError as e:
        # Handle ValidationError properly
        error_msg = str(e)
        if hasattr(e, 'messages'):
            error_msg = '\n'.join(e.messages)
        elif hasattr(e, 'message'):
            error_msg = e.message
        
        yield json.dumps({
            'type': 'error',
            'message': error_msg
        }) + '\n'
    except Exception as e:
        yield json.dumps({
            'type': 'error',
            'message': f'Terjadi kesalahan: {str(e)}'
        }) + '\n'
    
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_file_path):
            os.unlink(tmp_file_path)


@require_http_methods(["GET", "POST"])
@login_required(login_url='authentication:login')
def batch_import_students(request):
    """Batch import students from Excel file"""
    
    # Check if user is librarian
    if not _is_librarian(request.user):
        raise PermissionDenied("Hanya petugas perpustakaan yang dapat mengakses halaman ini.")
    
    if request.method == 'POST':
        form = StudentBatchImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                update_existing = form.cleaned_data.get('update_existing', False)
                
                # Save file temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    for chunk in excel_file.chunks():
                        tmp_file.write(chunk)
                    tmp_file_path = tmp_file.name
                
                # Return streaming response
                response = StreamingHttpResponse(
                    _process_import_stream(tmp_file_path, update_existing),
                    content_type='application/x-ndjson'
                )
                return response
            
            except Exception as e:
                return render(request, 'dropbox.html', {
                    'form': form,
                    'error': f"Terjadi kesalahan: {str(e)}"
                })
        
        else:
            # Form validation errors
            errors_list = []
            for field, errors_item in form.errors.items():
                for error in errors_item:
                    errors_list.append(str(error))
            
            return render(request, 'dropbox.html', {
                'form': form,
                'form_errors': errors_list
            })
    
    else:
        form = StudentBatchImportForm()
    
    context = {
        'form': form,
    }
    return render(request, 'dropbox.html', context)
