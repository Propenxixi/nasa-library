import re
import sys
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    import openpyxl
except ImportError:
    openpyxl = None


def clean_isbn(raw) -> str | None:
    """Bersihkan ISBN dari float / karakter non-numeric."""
    if raw is None:
        return None
    s = str(raw).strip()
    s = re.sub(r'\.0$', '', s)          # 9786234727210.0 → 9786234727210
    s = re.sub(r'[^0-9Xx]', '', s)      # buang non-digit kecuali X
    # Pastikan panjangnya masuk akal (ISBN-10 atau ISBN-13)
    if len(s) not in (10, 13):
        return s if s else None
    return s or None


def clean_str(val, maxlen=None) -> str:
    """Konversi nilai ke string bersih, truncate jika perlu."""
    if val is None:
        return ''
    s = str(val).strip()
    if maxlen:
        s = s[:maxlen]
    return s


def clean_int(val) -> int | None:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


class Command(BaseCommand):
    help = 'Import data buku dari file Excel (.xlsx) ke database'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Path ke file Excel (Dataset_Buku.xlsx)',
        )
        parser.add_argument(
            '--update',
            action='store_true',
            default=False,
            help='Update buku yang sudah ada (default: skip)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            default=False,
            help='Simulasi tanpa menyimpan ke database',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl belum terinstall. Jalankan: pip install openpyxl')

        from book.models import Book

        filepath  = options['excel_file']
        do_update = options['update']
        dry_run   = options['dry_run']

        self.stdout.write(f'\n📂 Membaca file: {filepath}')
        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  DRY RUN — tidak ada yang disimpan\n'))

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File tidak ditemukan: {filepath}')

        total_created = 0
        total_updated = 0
        total_skipped = 0
        total_error   = 0

        for sheet_name in wb.sheetnames:
            ws   = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Deteksi posisi kolom dari header (case-insensitive)
            headers = [str(h).lower().strip() if h else '' for h in rows[0]]

            def col(keywords):
                """Cari index kolom berdasarkan keyword."""
                for kw in keywords:
                    for i, h in enumerate(headers):
                        if kw in h:
                            return i
                return None

            idx_title   = col(['title'])
            idx_author  = col(['author'])
            idx_isbn    = col(['isbn'])
            idx_pages   = col(['length', 'pages', 'halaman'])
            idx_lang    = col(['language', 'bahasa'])
            idx_copies  = col(['copies', 'stok', 'total'])
            idx_shelf   = col(['shelf', 'rak', 'location'])
            idx_cover   = col(['image', 'cover'])
            idx_cat     = col(['category', 'kategori', 'genre', 'jenis'])

            if idx_title is None or idx_author is None:
                self.stdout.write(self.style.WARNING(f'  Sheet "{sheet_name}": kolom Title/Author tidak ditemukan, skip.'))
                continue

            data_rows = rows[1:]
            self.stdout.write(f'\n📚 Sheet: {sheet_name}  ({len(data_rows)} baris)')
            self.stdout.write('─' * 50)

            sheet_created = 0
            sheet_skipped = 0
            sheet_updated = 0
            sheet_error   = 0

            with transaction.atomic():
                for i, row in enumerate(data_rows):
                    def get(idx):
                        if idx is None or idx >= len(row):
                            return None
                        return row[idx]

                    title  = clean_str(get(idx_title), 500)
                    author = clean_str(get(idx_author), 500)
                    isbn   = clean_isbn(get(idx_isbn))

                    # Skip baris kosong
                    if not title:
                        continue

                    # Generate dummy ISBN jika tidak ada / tidak valid
                    if not isbn:
                        isbn = f'NOISBN-{sheet_name[:4].upper()}-{i+1:04d}'

                    pages      = clean_int(get(idx_pages))
                    language   = clean_str(get(idx_lang), 100) or 'Indonesian'
                    copies     = clean_int(get(idx_copies)) or 1
                    shelf      = clean_str(get(idx_shelf), 100)
                    cover_url  = clean_str(get(idx_cover), 1000) or None
                    category   = clean_str(get(idx_cat), 300) if idx_cat is not None else None

                    # Truncate shelf jika integer (beberapa sheet simpan angka)
                    if shelf and shelf.replace('.', '').isdigit():
                        shelf = f'Rak {int(float(shelf))}'

                    try:
                        existing = Book.objects.filter(isbn=isbn).first()

                        if existing:
                            if do_update:
                                existing.title         = title
                                existing.author        = author
                                existing.pages         = pages
                                existing.language      = language
                                existing.total_copies  = copies
                                existing.shelf_location= shelf
                                if cover_url:
                                    existing.cover_url = cover_url
                                if category:
                                    existing.category = category
                                if not dry_run:
                                    existing.save()
                                sheet_updated += 1
                                total_updated += 1
                            else:
                                sheet_skipped += 1
                                total_skipped += 1
                            continue

                        book = Book(
                            title          = title,
                            author         = author,
                            isbn           = isbn,
                            pages          = pages,
                            language       = language,
                            total_copies   = copies,
                            shelf_location = shelf,
                            cover_url      = cover_url or None,
                            category       = category,
                            status         = 'tersedia',
                        )
                        if not dry_run:
                            book.save()
                        sheet_created += 1
                        total_created += 1

                    except Exception as e:
                        self.stdout.write(self.style.ERROR(f'  ❌ Row {i+2}: {e} | title={title[:40]}'))
                        sheet_error   += 1
                        total_error   += 1

            self.stdout.write(
                f'  ✅ Dibuat: {sheet_created}  '
                f'🔄 Diupdate: {sheet_updated}  '
                f'⏭  Skip: {sheet_skipped}  '
                f'❌ Error: {sheet_error}'
            )

        wb.close()

        self.stdout.write('\n' + '═' * 50)
        self.stdout.write(self.style.SUCCESS(
            f'🎉 SELESAI\n'
            f'   Total dibuat  : {total_created}\n'
            f'   Total diupdate: {total_updated}\n'
            f'   Total skip    : {total_skipped}\n'
            f'   Total error   : {total_error}\n'
        ))

        if dry_run:
            self.stdout.write(self.style.WARNING('(Dry run — tidak ada perubahan ke DB)'))