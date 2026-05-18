import json
from datetime import datetime, timedelta
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Sum, Prefetch
from authentication.models import UserProfile
from attendance.models import Attendance
from book_loan.models import Loan
from book.models import Book
import openpyxl
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from django.conf import settings
import os

def is_authorized(user):
    """Check if user is a teacher or librarian"""
    if user.is_superuser:
        return True
    try:
        profile = user.profile
        return profile.role in ['teacher', 'librarian']
    except UserProfile.DoesNotExist:
        return False

@login_required
def report_page_view(request):
    if not is_authorized(request.user):
        messages.error(request, "Anda tidak memiliki akses ke halaman laporan.")
        return redirect('main:mainpage')
    
    return render(request, 'report/index.html')

def get_period_range(period, start_custom=None, end_custom=None):
    now = timezone.now()
    today = now.date()
    
    if period == 'custom' and start_custom and end_custom:
        try:
            start_date = datetime.strptime(start_custom, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_custom, '%Y-%m-%d').date()
            return start_date, end_date
        except (ValueError, TypeError):
            pass # Fallback to today if invalid

    if period == 'daily':
        start_date = today
        end_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=6)
        end_date = today
    elif period == 'monthly':
        start_date = today.replace(day=1)
        end_date = today
    elif period == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        start_date = today
        end_date = today
        
    return start_date, end_date

def get_attendance_data(start_date, end_date):
    # Fetch detailed attendance records in range
    records = Attendance.objects.filter(
        check_in_time__date__range=(start_date, end_date)
    ).select_related('user__profile').prefetch_related('activities').order_by('check_in_time')
    
    data = []
    for r in records:
        role = getattr(r.user.profile, 'role', '-') if hasattr(r.user, 'profile') else '-'
        role_display = 'Guru' if role == 'teacher' else 'Siswa' if role == 'student' else role.capitalize()
        
        # Format activities
        acts = [a.name for a in r.activities.all()]
        if r.custom_activity:
            acts.append(r.custom_activity)
        activities_str = ", ".join(acts) if acts else "-"
        
        data.append({
            'date': r.check_in_time.strftime('%d %b %Y').lstrip('0'),
            'nama': f"{r.user.first_name} {r.user.last_name}".strip() or r.user.username,
            'role': role_display,
            'aktivitas': activities_str
        })
    
    # Calculate summary metrics
    total_all = records.count()
    teacher_count = records.filter(user__profile__role='teacher').count()
    student_count = records.filter(user__profile__role='student').count()
    
    # Busiest Day
    date_counts = records.values('check_in_time__date').annotate(count=Count('id'))
    busiest_info = "—"
    if date_counts:
        busiest_day_record = max(date_counts, key=lambda x: x['count'])
        busiest_date = busiest_day_record['check_in_time__date']
        # Map weekday to Indonesian name as per Figma
        days_id = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        busiest_info = days_id[busiest_date.weekday()]

    summary = {
        'total': total_all,
        'teacher_count': teacher_count,
        'student_count': student_count,
        'busiest_day': busiest_info
    }
    return data, summary

def get_borrowing_data(start_date, end_date):
    today = timezone.now().date()
    loans_in_period = Loan.objects.filter(loan_date__date__range=(start_date, end_date))
    
    # 'Borrowed' here means currently active and NOT overdue
    borrowed = loans_in_period.filter(status='sedang_dipinjam', due_date__gte=today).count()
    returned = loans_in_period.filter(status='dikembalikan').count()
    
    # Overdue is either explicit status or active loan with past due date
    overdue = loans_in_period.filter(
        Q(status='terlambat') | 
        Q(is_overdue=True) | 
        Q(status='sedang_dipinjam', due_date__lt=today)
    ).count()
    
    table_loans = loans_in_period.select_related('user__profile', 'book').order_by('-loan_date')
    
    data = []
    for l in table_loans:
        st = l.status
        if st == 'dikembalikan':
            display_status = 'Dikembalikan'
        elif st == 'terlambat' or l.is_overdue or (st == 'sedang_dipinjam' and l.due_date and l.due_date < today):
            display_status = 'Terlambat'
        else:
            display_status = 'Dipinjam'
            
        data.append({
            'tanggal': l.loan_date.strftime('%d %b %Y').lstrip('0'),
            'judul_buku': l.book.title,
            'peminjam': f"{l.user.first_name} {l.user.last_name}".strip() or l.user.username,
            'kelas': getattr(l.user.profile, 'kelas', '-') if hasattr(l.user, 'profile') and l.user.profile.kelas else '-',
            'status': display_status,
            'tanggal_kembali': l.due_date.strftime('%d %b %Y').lstrip('0') if l.due_date else '-'
        })

    summary = {
        'total_dipinjam': borrowed,
        'total_dikembalikan': returned,
        'total_terlambat': overdue
    }
    return data, summary

def get_collection_data(start_date=None, end_date=None):
    # Fetch books with active loans pre-fetched to avoid N+1 queries
    active_loans_qs = Loan.objects.filter(status='sedang_dipinjam')
    books = Book.objects.all().prefetch_related(
        Prefetch('loans', queryset=active_loans_qs, to_attr='active_loans')
    )
    
    category_stats = {}
    total_physical_books = Book.objects.aggregate(total=Sum('total_copies'))['total'] or 0
    
    for book in books:
        # Resolve category
        raw_cat = book.category if book.category else 'Uncategorized'
        
        # Split by comma and strip whitespace
        cats = [c.strip() for c in raw_cat.split(',')]
        # Filter out empty strings
        cats = [c for c in cats if c]
        if not cats:
            cats = ['Uncategorized']
            
        # Calculate book availability stats once per book
        borrowed = len(book.active_loans)
        good_copies = book.total_copies - book.damaged_copies - book.lost_copies
        available = max(good_copies - borrowed, 0)
        
        for cat in cats:
            if cat not in category_stats:
                category_stats[cat] = {
                    'total_books': 0,
                    'tersedia': 0,
                    'dipinjam': 0
                }
            
            # Count the book's total copies and availability for EACH of its categories
            category_stats[cat]['total_books'] += book.total_copies
            category_stats[cat]['tersedia'] += available
            category_stats[cat]['dipinjam'] += borrowed
            
    # Convert aggregated dictionary back to the list format expected by the frontend
    data = []
    for cat, stats in category_stats.items():
        tot = stats['total_books']
        avl = stats['tersedia']
        brw = stats['dipinjam']
        
        persen = 0
        if tot > 0:
            persen = int((avl / tot) * 100)
            
        data.append({
            'category': cat,
            'total_books': tot,
            'tersedia': avl,
            'dipinjam': brw,
            'persentase_tersedia': f"{persen}%"
        })
        
    # Sort categories by total count descending
    data.sort(key=lambda x: x['total_books'], reverse=True)
        
    new_books_count = 0
    if start_date and end_date:
        res = Book.objects.filter(created_at__date__range=(start_date, end_date)).aggregate(Sum('total_copies'))
        new_books_count = res['total_copies__sum'] or 0

    summary = {
        'total': total_physical_books,
        'total_categories': len(category_stats),
        'new_books': new_books_count
    }
    return data, summary

@login_required
def api_report_preview(request):
    if not is_authorized(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    report_type = request.GET.get('type')
    period = request.GET.get('period')
    start_custom = request.GET.get('start_date')
    end_custom = request.GET.get('end_date')
    
    if not report_type or not period:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    start_date, end_date = get_period_range(period, start_custom, end_custom)
    
    if report_type == 'attendance':
        data, summary = get_attendance_data(start_date, end_date)
        columns = [
            {'header': 'Tanggal', 'key': 'date'},
            {'header': 'Nama', 'key': 'nama'},
            {'header': 'Role', 'key': 'role'},
            {'header': 'Aktivitas', 'key': 'aktivitas'}
        ]
    elif report_type == 'borrowing':
        data, summary = get_borrowing_data(start_date, end_date)
        columns = [
            {'header': 'TANGGAL', 'key': 'tanggal'},
            {'header': 'JUDUL BUKU', 'key': 'judul_buku'},
            {'header': 'PEMINJAM', 'key': 'peminjam'},
            {'header': 'KELAS', 'key': 'kelas'},
            {'header': 'STATUS', 'key': 'status'},
            {'header': 'TANGGAL KEMBALI', 'key': 'tanggal_kembali'}
        ]
    elif report_type == 'collection':
        data, summary = get_collection_data(start_date, end_date)
        columns = [
            {'header': 'Kategori', 'key': 'category'},
            {'header': 'Total Buku', 'key': 'total_books'},
            {'header': 'Tersedia', 'key': 'tersedia'},
            {'header': 'Dipinjam', 'key': 'dipinjam'},
            {'header': 'Persentase Tersedia', 'key': 'persentase_tersedia'}
        ]
    else:
        return JsonResponse({'error': 'Invalid report type'}, status=400)
    
    return JsonResponse({
        'status': 'success',
        'data': data,
        'summary': summary,
        'columns': columns,
        'period_label': f"{start_date} s/d {end_date}"
    })

@login_required
def api_report_export(request):
    if not is_authorized(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    report_type = request.POST.get('type')
    period = request.POST.get('period')
    export_format = request.POST.get('format')
    start_custom = request.POST.get('start_date')
    end_custom = request.POST.get('end_date')
    
    if not report_type or not period or not export_format:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    start_date, end_date = get_period_range(period, start_custom, end_custom)
    
    if report_type == 'attendance':
        data, _ = get_attendance_data(start_date, end_date)
        title = "Laporan Kehadiran Perpustakaan"
        headers = ['Tanggal', 'Nama', 'Role', 'Aktivitas']
        rows = [[item['date'], item['nama'], item['role'], item['aktivitas']] for item in data]
    elif report_type == 'borrowing':
        data, _ = get_borrowing_data(start_date, end_date)
        title = "Laporan Aktivitas Peminjaman"
        headers = ['TANGGAL', 'JUDUL BUKU', 'PEMINJAM', 'KELAS', 'STATUS', 'TANGGAL KEMBALI']
        rows = [[item['tanggal'], item['judul_buku'], item['peminjam'], item['kelas'], item['status'], item['tanggal_kembali']] for item in data]
    elif report_type == 'collection':
        data, _ = get_collection_data(start_date, end_date)
        title = "Laporan Koleksi Buku"
        headers = ['Kategori', 'Total Buku', 'Tersedia', 'Dipinjam', 'Persentase Tersedia']
        rows = [[item['category'], item['total_books'], item['tersedia'], item['dipinjam'], item['persentase_tersedia']] for item in data]
    else:
        return JsonResponse({'error': 'Invalid report type'}, status=400)

    period_str = f"{start_date} s/d {end_date}"

    # --- Summary Row Calculation (Unified for Excel and PDF) ---
    if report_type == 'attendance':
        t_siswa = sum(1 for item in data if item['role'] == 'Siswa')
        t_guru = sum(1 for item in data if item['role'] == 'Guru')
        t_total = len(data)
        rows.append(['TOTAL SISWA', '', '', t_siswa])
        rows.append(['TOTAL GURU', '', '', t_guru])
        rows.append(['TOTAL KESELURUHAN', '', '', t_total])
    elif report_type == 'borrowing':
        t_dipinjam = sum(1 for item in data if item['status'] == 'Dipinjam')
        t_terlambat = sum(1 for item in data if item['status'] == 'Terlambat')
        rows.append(['TOTAL DIPINJAM', '', '', '', t_dipinjam, ''])
        rows.append(['TOTAL TERLAMBAT', '', '', '', t_terlambat, ''])
    elif report_type == 'collection':
        t_books = sum(int(row[1]) for row in rows)
        t_avl = sum(int(row[2]) for row in rows)
        t_brw = sum(int(row[3]) for row in rows)
        rows.append(['TOTAL', t_books, t_avl, t_brw, '-'])

    if export_format == 'excel':
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        num_cols = len(headers)
        
        # --- Formal Letterhead Image ---
        img_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'kop.png')
        if os.path.exists(img_path):
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(img_path)
            # Set scale to fit width (approx 7.5 inches / 19cm)
            img.width = 700 
            img.height = 140
            ws.add_image(img, 'A1')
            # Merge cells for the header
            ws.merge_cells(start_row=1, start_column=1, end_row=7, end_column=num_cols)
            # Add spacers to clear the merged area
            for _ in range(8): ws.append([])
        else:
            # Fallback to text
            def add_merge_header(row, text, size, bold=True):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
                cell = ws.cell(row=row, column=1)
                cell.value = text
                cell.font = Font(name='Arial', size=size, bold=bold)
                cell.alignment = Alignment(horizontal='center')
            add_merge_header(1, "PEMERINTAH PROVINSI D.K.I JAKARTA", 12)
            add_merge_header(2, "SMA NEGERI 61 JAKARTA", 14)
            add_merge_header(3, "Jl. Taruna Jl. Pahlawan Revolusi, Pd. Bambu, Kec. Duren Sawit, Jakarta Timur", 10, False)
            for _ in range(5): ws.append([])
        
        # Report Title Row (shifted to row 9 to give a small gap)
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=num_cols)
        ws['A9'] = title.upper()
        ws['A9'].font = Font(name='Arial', size=14, bold=True)
        ws['A9'].alignment = Alignment(horizontal='center')
        
        ws.append([]) # spacer
        ws.append([f"Periode: {period_str}"])
        ws.append([f"Tanggal Cetak: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"])
        # Table Headers
        ws.append([]) # spacer
        header_row_idx = ws.max_row + 1
        ws.append(headers)
        
        # Use a more standard grey hex with FF alpha prefix
        header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[header_row_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = header_border
            
        # Data
        for row in rows:
            ws.append(row)
            
        # Borders and Total Row Styling
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for r_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell_val = str(ws.cell(row=r_idx, column=1).value or "")
            is_total = 'TOTAL' in cell_val
            for cell in ws[r_idx]:
                cell.border = thin_border
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-adjust column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in col:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjusted_width = (max_length + 4) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 60)

        output = BytesIO()
        wb.save(output)
        
        filename = f"report_{report_type}_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    elif export_format == 'pdf':
        buffer = BytesIO()
        class NumberedCanvas(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                canvas.Canvas.__init__(self, *args, **kwargs)
                self.pages = []
            def showPage(self):
                self.pages.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                page_count = len(self.pages)
                for page in self.pages:
                    self.__dict__.update(page)
                    self.draw_page_number(page_count)
                    canvas.Canvas.showPage(self)
                canvas.Canvas.save(self)
            def draw_page_number(self, page_count):
                self.setFont("Times-Roman", 10)
                self.drawRightString(200 * mm, 15 * mm, f"Halaman {self._pageNumber} dari {page_count}")
        
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=30*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontName='Times-Bold', fontSize=14, alignment=TA_CENTER, spaceAfter=2)
        address_style = ParagraphStyle('AddressStyle', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, alignment=TA_CENTER, spaceAfter=2)
        report_title_style = ParagraphStyle('ReportTitle', parent=styles['Normal'], fontName='Times-Bold', fontSize=16, alignment=TA_CENTER, spaceBefore=20, spaceAfter=25)
        info_style = ParagraphStyle('InfoStyle', parent=styles['Normal'], fontName='Times-Roman', fontSize=11, spaceAfter=4)

        logo_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'logoSMAN61.png')
        img = Image(logo_path, width=25*mm, height=25*mm) if os.path.exists(logo_path) else Paragraph("", styles['Normal'])

        header_text = [
            Paragraph("PEMERINTAH PROVINSI D.K.I JAKARTA", title_style),
            Paragraph("SMA NEGERI 61 JAKARTA", ParagraphStyle('SchoolName', parent=title_style, fontSize=16, spaceBefore=4, spaceAfter=8)),
            Paragraph("Jl. Taruna Jl. Pahlawan Revolusi, Pd. Bambu, Kec. Duren Sawit, Kota Jakarta Timur", address_style),
            Paragraph("Daerah Khusus Ibukota Jakarta 13430", address_style),
        ]

        header_table = Table([[img, header_text]], colWidths=[30*mm, 140*mm])
        header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (1, 0), (1, 0), 'CENTER'), ('LEFTPADDING', (1, 0), (1, 0), 10)]))
        elements.append(header_table)
        
        from reportlab.platypus import HRFlowable
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.black, spaceBefore=5, spaceAfter=1))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.black, spaceBefore=1, spaceAfter=10))

        elements.append(Paragraph(title.upper(), report_title_style))
        elements.append(Paragraph(f"Periode: {period_str}", info_style))
        elements.append(Paragraph(f"Tanggal Cetak: {datetime.now().strftime('%d %B %Y %H:%M:%S')}", info_style))
        elements.append(Spacer(1, 10))

        # --- Table Implementation with Wrapping ---
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName='Times-Roman', fontSize=10, leading=12)
        header_p_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontName='Times-Bold', fontSize=10, alignment=TA_CENTER)
        
        formatted_headers = [Paragraph(h, header_p_style) for h in headers]
        formatted_rows = []
        for row in rows:
            f_row = []
            for i, val in enumerate(row):
                # Apply wrapping to columns that might be long
                if i in [1, 3] or (report_type == 'borrowing' and i == 1):
                    f_row.append(Paragraph(str(val), cell_style))
                else:
                    f_row.append(val)
            formatted_rows.append(f_row)

        table_data = [formatted_headers] + formatted_rows
        
        if report_type == 'collection':
            col_widths = [55*mm, 25*mm, 25*mm, 25*mm, 40*mm]
        elif report_type == 'attendance':
            col_widths = [25*mm, 45*mm, 25*mm, 75*mm]
        else:
            col_widths = None

        t = Table(table_data, repeatRows=1, colWidths=col_widths)
        table_style = [
            ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]
        
        # Style TOTAL rows
        for idx, row in enumerate(table_data):
            if idx > 0 and 'TOTAL' in str(row[0]):
                table_style.append(('FONTNAME', (0, idx), (-1, idx), 'Times-Bold'))
                table_style.append(('BACKGROUND', (0, idx), (-1, idx), colors.lightgrey))
                # Adjust span based on column count
                span_end = 3 if report_type == 'borrowing' else 2
                table_style.append(('SPAN', (0, idx), (span_end, idx))) 
                table_style.append(('ALIGN', (0, idx), (0, idx), 'LEFT'))

        t.setStyle(TableStyle(table_style))
        elements.append(t)
        doc.build(elements, canvasmaker=NumberedCanvas)
        
        filename = f"report_{report_type}_{period}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return JsonResponse({'error': 'Invalid format'}, status=400)
